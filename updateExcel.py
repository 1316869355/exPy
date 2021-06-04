# -*- coding:utf-8 -*-
#! python3
# updateExcel.py - Corrects costs in produce sales spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('example_copy.xlsx')
sheet = wb.active

# The produce types and their updated prices
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery':1.19,
                 'Lemon': 1.27}

# TODO: Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(rowNum, 1).value
    print('ProduceName = '+ str(produceName))
    if produceName in PRICE_UPDATES:
        sheet.cell(rowNum, 2).value=PRICE_UPDATES[produceName]
        
wb.save('example_copy.xlsx')
